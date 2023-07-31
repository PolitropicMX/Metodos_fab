    # ZONA DE IMPORTACION DE CODIGO
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
import openpyxl
import tkinter
import os
import glob
from natsort import natsorted
from productos_reactivos import *



# COMENTARIOS ACERCA DEL CODIGO:
# SI SE AÑADE MAS FILAS EN LA PRIMERA PARTE, RECORDAR CAMBIAR TODOS LOS 'x#' EJEMPLO (A1,B2) DE TODAS LAS LINEAS QUE REFERENCIEN 


wb_test = Workbook()

class Save_module:
    def __init__(self):
        self.titulos = []# guarda los nombres de los archivos que ha creado
        self.wb_test = Workbook()# Libro de trabajo
        self.clase_productos = Productos()# IMPORTA LOS DATOS DE LOS DEMAS PRODUCTOS
        self.lista_productos = self.clase_productos.productos# IMPORTA EL DICCIONARIO
    def active_module(self):
        self.run = True
        while self.run:
            print('Bienvenido al sistema de gestión de metodos de fabricacion')
            print('Que desea Hacer')

            opciones = ['Crear un metodo de Fabricación (1)',
                        'Editar un metodo existente (2)',
                        'Mostrar Metodos (3)',
                        'Terminar programa (4)'
                        ]
            
            for opcion in opciones:
                print(opcion)

            eleccion = input('>>> ')

            if eleccion == '1': # Crear un metodo de Fabricación (1)
                
                print('Ha elegido crear un nuevo metodo')
                print('Como se llama el producto')
                
                archivo_nuevo = input('>>> ')
                
                wb_test.save(archivo_nuevo+'.xlsx')
                
                print(' SE HA GENERADO EL ARCHIVO')
                
                # SSE GUARDA EL NOMBRE DEL PRODUCTO EN UN ARCHIVO DE TEXTO
                # LLAMADO "NOMBRES.TXT"
                
                self.f = open("nombres.txt","a")
                self.f.write('<>')
                self.f.write(archivo_nuevo)
                self.f.close()
                # abrir de nuevo el archivo recien creado
                print('Se agregara los datos basicos al archivos')
                self.pegar_parte_superior(self.wb_test,archivo_nuevo)
                
            elif eleccion == '2':# PARA EDITAR ARCHIVOS
                print('Ha elegido modificar el metodo')
                self.mostrar_archivos()# Se muestran los archivos creados
                
            elif eleccion == '3':# UNICAMENTE MOSTRAR EL REGISTRO DE ARCHIVOS
                self.mostrar_archivos()# Se muestran los archivos creados
            elif eleccion == '4':# Terminar programa
                print('Programa terminado')
                self.run = False

            # Instruccion de termino del Loop
            if self.run == False:
                pass
            else:
                print('Desea terminar con el programa? SI(1), NO(0)')
                self.run = input('>>> ')
                if self.run == '1':
                    self.run = False
                    print('Programa terminado')
                else:
                    self.run = True

                    
    def mostrar_archivos(self):
        self.titulos = []# Se debe reinicar la lista de los titulos
        print('Se muestra la lista de archivos guardados')
        self.f = open("nombres.txt","r")
        self.strtitulos = str(self.f.read()) # String de titulos
        self.initiallimit = 2
        # Ciclo FOR que itera a traves del string colectando todos los titulos
        for i,j in enumerate(self.strtitulos):
            if j == '<' and not i == 0:
                self.titulos.append(self.strtitulos[self.initiallimit:i])
                self.initiallimit = i+2
            if i == len(self.strtitulos)-1:
                self.titulos.append(self.strtitulos[self.initiallimit:i+1])
        for i,j in enumerate(self.titulos):
            print(f'{i+1} archivo: {j}')

    def stylin(self,cell,f="Calibri",fc="000000",fs=11,fgc="ffffff",h="center",v="center",wt=True,bo=False):
        # PARAMETERS: CELL:'A2', FC = FONT COLOR, FS = FONT SIZE, FGC = FOREGROUND COLOR, , h = Horizontal, v = Vertical, wt = wrap_text
        self.ws[cell].font = Font(name=f,color=fc,size=fs,b=bo)
        self.ws[cell].fill = PatternFill("solid", fgColor=fgc)
        self.ws[cell].alignment = Alignment(horizontal=h, vertical=v,wrap_text= wt)

    def combine(self,cell1,cell2):
        self.ws.merge_cells(cell1+":"+cell2)
            
    def pegar_parte_superior(self,wb,nombre):
        self.wb = wb
        self.ws = self.wb.active
            # SE CAMBIAN EL TAMAÑO DE LAS COLUMNAS
            # por cada unidad en esta linea son 7 px
        self.ws.column_dimensions['A'].width = 21
        self.ws.column_dimensions['B'].width = 8
        self.ws.column_dimensions['C'].width = 8
        self.ws.column_dimensions['D'].width = 8
        self.ws.column_dimensions['E'].width = 22
        self.ws.column_dimensions['F'].width = 22
            # SE CAMBIAN EL TAMAÑO DE LAS FILAS
        self.ws.row_dimensions[3].height = 50
        self.ws.row_dimensions[4].height = 50
            #Obtener la fecha actual del sistema
        Fecha_elaboracion = str(date.today())
            # Recomencdaciones generales y 
        parte_superior = [
                   ['B1','Metodo de Fabricacion Estandar'],
                   ['F1','Documento Controlado'],
                   ['A2','PETRA'], # 0
                   ['B2', nombre], # 1
                   ['F2','Fecha de elaboracion '+Fecha_elaboracion], # 2
                   ['A4','Equipo a utilizar'],# 3
##                   ['B2','POLIPASTO, TINA CHICA DE ACERO INOX., DISPERSOR, BASCULAS, PLATAFORMA( GALLINA), PATIN'],# 4
                   ['E3','ANTES DE INICIAR, LIMPIE PERFECTAMENTE EL EQUIPO QUE VAYA A UTILIZAR, Y MANTENGA SU LUGAR DE TRABAJO LIMPIO, TENGA UNA FRANELA PARA EVITAR ENSUCIAR'],# 5
                   ['A3','Verifique que tenga todos los insumos indicados en la orden de producción, y que se encuentren debidamente identificados y APROBADOS'],# 6
                   ['E4','Si en su orden encuentra otro numero de inventario diferente al MFE, consulte con el SPR antes de iniciar su proceso	'],
                   ['A5','No. Inven y Descripcion'],
                   ['B5','PICTOGRAMAS'],
                   ['E5','INDICACION'],
                   ['F5','ANTES DE USAR...']
                   ]
            # COMBINAR CELDAS PARTE SUPERIOR
##        self.ws.merge_cells("B1:E1")
        self.combine("B1","E1")
        self.combine("B2","E2")
        self.combine("E3","F3")
        self.combine("A3","D3")
        self.combine("E4","F4")
        self.combine("B5","D5")
        self.combine("B4","D4")
##        self.ws.merge_cells("B2:D2")  ESTA NO, POR QUE UNA VEZ SELECCIONADO EL PRODUCTO SE AÑADIRÁ EL EQUIPO A  UTILIZAR

        
            # PICTOGRAMAS
        letra_picto = ['a','c','i','m','t']
            # TODAS LAS IMAGENES
        images = {'a':'C:/Users/Fernando.Lopez/Downloads/los_codigos/los_codigos/openpyxl/pruebas/a.png',# ATENCION
                  'c':'C:/Users/Fernando.Lopez/Downloads/los_codigos/los_codigos/openpyxl/pruebas/c.png',# CORROSIVO
                  'i':'C:/Users/Fernando.Lopez/Downloads/los_codigos/los_codigos/openpyxl/pruebas/i.png',# INFLAMABLE
                  'm':'C:/Users/Fernando.Lopez/Downloads/los_codigos/los_codigos/openpyxl/pruebas/m.png',# DAÑO AMBIENTAL
                  't':'C:/Users/Fernando.Lopez/Downloads/los_codigos/los_codigos/openpyxl/pruebas/t.png',# TOXICO
                  'b1':'C:/Users/Fernando.Lopez/Downloads/los_codigos/los_codigos/openpyxl/advertisement.png' # BANNER DE PREPARACION
                  }
        # Images = {'a':'Ruta de a',
        #           'c':'Ruta de c',...}
            # CICLO FOR: AQUI SE CARGA A "PARTE SUPERIOR" AL ARCHIVO
        for i,j in enumerate(parte_superior):
                # SE AÑADE EL TEXTO 
            self.ws[j[0]].value = j[1]
                # SE ENCUADRA EL TEXTO
            self.ws[j[0]].alignment = Alignment(wrap_text= True)

        # SE ESTILIZA LA PARTE DE ARRIBA
            # // COLORES VERDE CLARO 08bd83
            # // COLOR VERDE MAS CLARO 2cf2b3
            # // COLORES VERDE PETRA 046546

        
            # SE ESTILIZA EL TITULO
                # Metodo de Fabricacion Estandar
        self.stylin('A1',fgc="000000")
        self.stylin('B1',f='arial',fc="ffffff",fs=14,fgc="000000")
                # Documento Confidencial
        self.stylin('F1',f='arial',fc="ffffff",fs=10,fgc="000000")
                # PETRA
        self.stylin('A2',f="Berlin Sans FB Demi",fc="046546",fs=20,fgc="ffffff",)
                # Titulo del Documento 7 # Color PETRA 046546
        self.stylin('B2',f="Chicago",fc="ffffff", fs=20,fgc="046546")
                # Fecha de Elaboracion
        self.stylin('F2',f="Albertus MT")
            # Equipo a utilizar
        self.stylin('A3',fc="ffffff",fgc="08bd83")
            # ANTES DE INICIAR LIMPIE...
        self.stylin('E3',fc="ffffff", fs=9,fgc="08bd83" )
            # Verifique que tenga
        self.stylin('A4',fc="000000", fs=9,fgc="2cf2b3" )
            # No. Inven y Descripcion
        self.stylin('A5',fc="ffffff",fgc="000000" )
            # PICTOGRAMAS
        self.stylin('B5',fc="ffffff",fgc="000000" )
            # Indicacion
        self.stylin('E5',fc="ffffff",fgc="000000" )
            # Antes de usar
        self.stylin('F5',fc="ffffff",fgc="000000" )
        
        
        # AQUI YA SE CARGO LA INFORMACION DE LA PARTE SUPERIOR
        ############################################################################################################################################
        
        # A PARTIR DE AQUI SE CARGA LA INFORMACIÓN DE LO REACTIVOS
        
            # DESEA CARGAR LA INFORMACION?
        print('DESEA CARGAR LA INFORMACION?')
        op1 = input('>>> ')

            # SI DESEO CARGAR LA INFORMACION
        if op1 == 'si':
                # SE PREGUNTA AL USUARIO
            print('Se cargaran los paquetes de datos')
            print('CUAL DESEA AÑADIR?')
                # SE ITERA A TRAVES DE DICCIONARIO 'PRODUCTOS Y REACTIVOS' DE LA BASE DE DATOS 
            for i,producto in enumerate(self.lista_productos): 
                print(f'{i}: {producto}')
            print('elija segun su NOMBRE:')
            op2 = input('>>> ')
                # SE SELECCIONA EL PAQUETE DE  INFORMACIÓN
            if op2 in self.lista_productos:# SE COMPRUEBA SI EL STRING INGRESADO ESTA EN EL DICCIONARIO
                self.reactivos_hoja = []
                    # FILA DONDE SE EMPIEZA A AGREGAR LA INFORMACION
                fila = 6
                print('Si esta')
                    # SE ACOMODAN LOS DATOS
                    # EQUIPO A UTILIZAR
                equipo_a_usar = self.lista_productos[op2]["Equipo a utilizar"]["Equipo a utilizar"]
                self.ws['B4'].value = equipo_a_usar
                self.ws.merge_cells("B3:D3")
                self.ws['B4'].alignment = Alignment(horizontal="center", vertical="center",wrap_text= True)
                self.envase_producto = self.lista_productos[op2]["Envase"]["Envase_producto"]
                self.envase_control = self.lista_productos[op2]["Envase"]["Envase_control"]
                    # Se itera a traves de el producto seleccionado
                    # reactivo es la llave del nivel 2
                    # {key1: value1}
                    # {Key1: {Key2:Value2}}
                    # {Key1: {Key2: {Key3: value}}}
                n_paso = 1
                n_banner = 1
                for i,reactivo in enumerate(self.lista_productos[str(op2)].keys()):
                    print(reactivo)
                    
                    # AQUI ES CUANDO SE CARGA LA INFORMACION DE LOS REACTIVOS
                        # NO SE ELIGE EL ULTIMO NI EL PENULTIMO ELEMENTO DE CADA PRODUCTO
                    if i != len(self.lista_productos[str(op2)].keys())-1 and i != len(self.lista_productos[str(op2)].keys())-2:
                            # SE MUESTRA EL NUMERO DE ITERACION

                        # Es reactivo o paso?
                        if reactivo[:4] == "Paso":# LOS PRIMEROS 4 CARACTERES
                            print(f'{i}# Iteracion Paso')
                            paso = self.lista_productos[op2][f"Paso {n_paso}"]["texto"]
                            self.reactivos_hoja.append(['A'+str(fila),paso])
                            self.ws.merge_cells('A'+str(fila)+':F'+str(fila))
                            n_paso += 1
                            fila += 1
                        elif reactivo[:6] == "Banner":
                            ruta = images[self.lista_productos[op2][f'Banner {n_banner}']]
                            self.ws.add_image(Image(ruta),anchor='A'+str(fila))
                            n_banner += 1
                            fila += 3
                        else:
                            print(f'{i}# Iteracion ')

                                # SE AÑADE LA INFORMACION DE LA BASE DE DATOS AL ARRAY
                                    # No inventario
                            self.reactivos_hoja.append(['A'+str(fila),reactivo])
                                    # DESCRIPCION
                            descripcion = self.lista_productos[op2][reactivo]['descripcion']
                            self.reactivos_hoja.append(['A'+str(fila+1),descripcion])
                                    # INDICACION
                            indicacion = self.lista_productos[op2][reactivo]['indicacion']
                            self.reactivos_hoja.append(['E'+str(fila),indicacion])
                                    # REVISION DE USO
                            revision = self.lista_productos[op2][reactivo]['revision']
                            self.reactivos_hoja.append(['F'+str(fila),revision])
                                # SE MODIFICAN LAS CELDAS (ESTILOS)
                            self.ws.merge_cells('A'+str(fila+1)+':A'+str(fila+2))# se combinan (DESCRIPCION)
                            self.ws.merge_cells('E'+str(fila)+':E'+str(fila+2))# se combinan (INDICACION)
                            self.ws.merge_cells('F'+str(fila)+':F'+str(fila+2))# se combinan (REVISION)
                            self.ws.row_dimensions[fila+1].height = 30

                                # SE AÑADEN LOS PICTOGRAMAS
                            peligro = self.lista_productos[op2][reactivo]['peligro']
                            if peligro != '0':
                                for k,pictograma in enumerate(peligro):
                                    ruta = images[pictograma]
##                                    print(ruta)
                                    self.ws.add_image(Image(ruta),anchor=chr(66+k)+str(fila))
                                    
                                    if pictograma == 'm':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Daño Ambiental'
                                    elif pictograma == 'a':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Atencion'
                                    elif pictograma == 'c':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Corrosivo'
                                    elif pictograma == 't':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Toxico'
                                    elif pictograma == 'i':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Inflamable'
                                        
                                    self.stylin(chr(66+k)+str(fila+2),fs=8)
                                    self.stylin(chr(66)+str(fila),fgc="046546")
                                    self.stylin(chr(66)+str(fila+1),fgc="046546")
                                    self.stylin(chr(67)+str(fila),fgc="046546")
                                    self.stylin(chr(67)+str(fila+1),fgc="046546")
                                    self.stylin(chr(68)+str(fila),fgc="046546")
                                    self.stylin(chr(68)+str(fila+1),fgc="046546")
                                    
                                    
                            fila += 3
                    # SE AÑADEN AL EXCEL LOS REACTIVOS
                solo_reac = 1
                for i,anadir in enumerate(self.reactivos_hoja):
                    self.ws[anadir[0]].value = anadir[1]
                    self.ws[anadir[0]].alignment = Alignment(horizontal="center", vertical="center", wrap_text= True)# SE ENCUADRA EL TEXTO
                        # EL NOMBRE DEL REACTIVO SE RESALTA

                    if anadir[0][0] == 'A' and solo_reac == 1 and (anadir[1][0] == '1' or anadir[1][0] == '2'):
##                        print(f' funciono ? {anadir[0]}')
                        self.ws[anadir[0]].font = Font(name="Albertus MT",color="ffffff", size=12)
                        self.ws[anadir[0]].fill = PatternFill("solid", fgColor="046546")
                        solo_reac = 3
##                    elif anadir[1][0] != '1' or :
##                        self.ws.row_dimensions[int(anadir[0][1])].height = 40
##                        solo_reac = 1
                    else:
                        solo_reac = 1
                    if anadir[0][0] == 'A' and  not anadir[1][0] == '1':
                        print(anadir[0][1])
                        if len(anadir[1]) >= 85:
                            self.ws.row_dimensions[int(anadir[0][1:])].height = 40
                        elif len(anadir[1]) >= 160:
                            self.ws.row_dimensions[int(anadir[0][1:])].height = 40
                            self.ws[anadir[0]].font(size=10)
##                    print(f'anadir es {anadir[0]} y su valor {anadir[1]}')
                    
            ####################################################################################################################################################3

                    # INFORMACION A AÑADIR: PARTE INFERIOR
                parte_inferior = [['A'+str(fila),'TOMA DE MUESTRA'],
                                  ['A'+str(fila+1),f'1)Tome una muestra en un envase de {self.envase_control} (Debe estar limpio y seco)'],
                                  ['A'+str(fila+2),'2) Con un marcador, coloque el nombre del producto.'],
                                  ['A'+str(fila+3),'3) Entregue la muestra al ICC junto con la orden de producción correspondiente.'],
                                  ['D'+str(fila),'ENVASADO'],
                                  ['D'+str(fila+1),'1) Verifique que la malla este limpia y no tenga residuos de otro material.'],
                                  ['D'+str(fila+2),'2) Verifique que los envases no esten sucios y no tengan polvo.'],
                                  ['D'+str(fila+3),'3) Al cerrar los envases verifique que esten bien cerrados.'],
                                  ['F'+str(fila+1),'Una vez aprobado su producto proceda a envasar de acuerdo a lo indicado en la orden de producción. Envase filtrando su material con una malla de 150 micras y envase con agitación'],
                                  ['A'+str(fila+5),'ELABORÓ'],
                                  ['A'+str(fila+6),'REVISÓ'],
                                  ['A'+str(fila+7),'AUTORIZÓ'],
                                  ['B'+str(fila+5),'Fernando López V.'],
                                  ['B'+str(fila+6),'Leonardo Sánchez M.'],
                                  ['B'+str(fila+7),'José de Jesus Cruz M.'],
                                  ['E'+str(fila+5),'Diseño y Desarrollo'],
                                  ['E'+str(fila+6),'Jefe de Producción'],
                                  ['E'+str(fila+7),'Jefe de Laboratorio'],
                                  ['F'+str(fila+4),'Firma']
                                  
                                  ]
                    # SE ESTILIZA LA HOJA: TAMAÑO DE LAS FILAS
                self.ws.row_dimensions[fila+1].height = 30
                self.ws.row_dimensions[fila+2].height = 30
                self.ws.row_dimensions[fila+3].height = 30
                self.ws.row_dimensions[fila+5].height = 25
                self.ws.row_dimensions[fila+6].height = 25
                self.ws.row_dimensions[fila+7].height = 25
                    # SE PEGA LA INFORMACION
                for i,j in enumerate(parte_inferior):
                    self.ws[j[0]].value = j[1] # SE AÑADE EL TEXTO 
                    self.ws[j[0]].alignment = Alignment(wrap_text= True)# SE ENCUADRA EL TEXTO
                    if j[0][0] == 'A' and int(j[0][1]) < fila:
                        self.ws[j[0]].fill = PatternFill("solid", fgColor="08bd83")
                    elif j[0][0] == 'D':
                        self.ws[j[0]].fill = PatternFill("solid", fgColor="2cf2b3")
                    # SE COMBINAN LAS CELDAS
                self.ws.merge_cells("A"+str(fila)+":C"+str(fila))
                self.ws.merge_cells("A"+str(fila+1)+":C"+str(fila+1))
                self.ws.merge_cells("A"+str(fila+2)+":C"+str(fila+2))
                self.ws.merge_cells("A"+str(fila+3)+":C"+str(fila+3))
                self.ws.merge_cells("D"+str(fila)+":F"+str(fila))
                self.ws.merge_cells("D"+str(fila+1)+":E"+str(fila+1))
                self.ws.merge_cells("D"+str(fila+2)+":E"+str(fila+2))
                self.ws.merge_cells("D"+str(fila+3)+":E"+str(fila+3))
                self.ws.merge_cells("F"+str(fila+1)+":F"+str(fila+3))
                self.ws.merge_cells("B"+str(fila+5)+":D"+str(fila+5))
                self.ws.merge_cells("B"+str(fila+6)+":D"+str(fila+6))
                self.ws.merge_cells("B"+str(fila+7)+":D"+str(fila+7))
                    # ESTILO DE CELDA
                self.stylin('A'+str(fila),f=="Albertus MT",fc="ffffff",fs=14,fgc="033827")
                self.stylin('D'+str(fila),f="Albertus MT",fc="ffffff",fs=14,fgc="033827")
                self.ws['F'+str(fila+5)].border = Border(bottom=Side(style='medium'))
                self.ws['F'+str(fila+6)].border = Border(bottom=Side(style='medium'))
                self.ws['F'+str(fila+7)].border = Border(bottom=Side(style='medium'))
                self.stylin('A'+str(fila),f=="Albertus MT",fc="ffffff",fs=14,fgc="033827",bo = True)
                
                
                self.reactivos_hoja =  []
            else:# EL USUARIO SE EQUIVOCO Y NO ESTA EN EL DICCIONARIO
                print('no está')

            # AQUI SE CARGA LA ULTIMA PARTE DEL PROGRAMA
            
            
        else: # NO DESEO CARGAR LA INFORMACION
            pass
        print(f'ULTIMA FILA {fila}')


##      ULTIMA INSTRUCCION ANTES DE CERRAR EL ARCHIVO
        self.wb.save(f'{nombre}.xlsx')

        

### INICIALIZAMOS EL PROGRAMA ###
clase_save = Save_module()
clase_save.active_module()

