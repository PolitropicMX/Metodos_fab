nfrom datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import openpyxl
import tkinter
import os
import glob
from natsort import natsorted
from productos_reactivos import *

# RETO: CREA 3 ARCHIVOS XLSX Y LOS NOMBRES DE UNA LISTA
nombres = ['Producto 1','Producto 2','Producto 3']
wb_test = Workbook()
# SE COMPLETÓ EL RETO
##for file in nombres:
##    wb_test.save(file+'.xlsx')

class Save_module:
    def __init__(self):
        self.titulos = []# guarda los nombres de los archivos que ha creado
        self.wb_test = Workbook()# Libro de trabajo
        self.clase_productos = Productos()# IMPORTA LOS DATOS DE LOS DEMAS PRODUCTOS
        self.lista_productos = self.clase_productos.productos# IMPORTA EL DICCIONARIO
    def active_module(self):
        self.run = True
        while self.run:
            print('Bienvenido al sistema de gestión de metodos de fabricacion')
            print('Que desea Hacer')

            opciones = ['Crear un metodo de Fabricación (1)',
                        'Editar un metodo existente (2)',
                        'Mostrar Metodos (3)',
                        'Terminar programa (4)'
                        ]
            
            for opcion in opciones:
                print(opcion)

            eleccion = input('>>> ')

            if eleccion == '1': # Crear un metodo de Fabricación (1)
                
                print('Ha elegido crear un nuevo metodo')
                print('Como se llama el producto')
                
                archivo_nuevo = input('>>> ')
                
                wb_test.save(archivo_nuevo+'.xlsx')
                
                print(' SE HA GENERADO EL ARCHIVO')
                
                # SSE GUARDA EL NOMBRE DEL PRODUCTO EN UN ARCHIVO DE TEXTO
                # LLAMADO "NOMBRES.TXT"
                
                self.f = open("nombres.txt","a")
                self.f.write('<>')
                self.f.write(archivo_nuevo)
                self.f.close()
                # abrir de nuevo el archivo recien creado
                print('Se agregara los datos basicos al archivos')
                self.pegar_parte_superior(self.wb_test,archivo_nuevo)
                
            elif eleccion == '2':# PARA EDITAR ARCHIVOS
                print('Ha elegido modificar el metodo')
                self.mostrar_archivos()# Se muestran los archivos creados
                
            elif eleccion == '3':# UNICAMENTE MOSTRAR EL REGISTRO DE ARCHIVOS
                self.mostrar_archivos()# Se muestran los archivos creados
            elif eleccion == '4':# Terminar programa
                print('Programa terminado')
                self.run = False

            # Instruccion de termino del Loop
            if self.run == False:
                pass
            else:
                print('Desea terminar con el programa? SI(1), NO(0)')
                self.run = input('>>> ')
                if self.run == '1':
                    self.run = False
                    print('Programa terminado')
                else:
                    self.run = True

                    
    def mostrar_archivos(self):
        self.titulos = []# Se debe reinicar la lista de los titulos
        print('Se muestra la lista de archivos guardados')
        self.f = open("nombres.txt","r")
        self.strtitulos = str(self.f.read()) # String de titulos
        self.initiallimit = 2
        # Ciclo FOR que itera a traves del string colectando todos los titulos
        for i,j in enumerate(self.strtitulos):
            if j == '<' and not i == 0:
                self.titulos.append(self.strtitulos[self.initiallimit:i])
                self.initiallimit = i+2
            if i == len(self.strtitulos)-1:
                self.titulos.append(self.strtitulos[self.initiallimit:i+1])
        for i,j in enumerate(self.titulos):
            print(f'{i+1} archivo: {j}')


            
    def pegar_parte_superior(self,wb,nombre):
        self.wb = wb
        self.ws = self.wb.active
        
        self.ws.column_dimensions['A'].width = 20# por cada unidad en esta linea son 7 px
        self.ws.column_dimensions['B'].width = 27# por cada unidad en esta linea son 7 px
        self.ws.column_dimensions['C'].width = 27# por cada unidad en esta linea son 7 px
        self.ws.column_dimensions['D'].width = 20# por cada unidad en esta linea son 7 px
        self.ws.row_dimensions[3].height = 60
        #Obtener la fecha actual del sistema
        Fecha_elaboracion = str(date.today())
        # Recomencdaciones generales y 
        parte_superior = [['A1','PETRA'], # 0
                   ['B1', nombre], # 1
                   ['D1','Fecha de elaboracion '+Fecha_elaboracion], # 2
                   ['A2','Equipo a utilizar'],# 3
                   ['B2','POLIPASTO, TINA CHICA DE ACERO INOX., DISPERSOR, BASCULAS, PlATAFORMA( GALLINA), PATIN'],# 4
                   ['C2','ANTES DE INICIAR LIMPIE PERFECTAMENTE EL EQUIPO QUE VAYA A UTILIZAR, Y MANTENGAN SU LUGAR DE TRABAJO LIMPIO, TENGA UNA FRANELA PARA EVITAR ENSUCIAR'],# 5
                   ['A3','Verifique que tenga todos los insumos indicados en la orden de producción, y que se encuentren debidamente identificados y APROBADOS'],# 6
                   ['C3','Si en su orden encuentra otro numero de inventario diferente al MFE, consulte con el SPR antes de iniciar su proceso	'],
                   ['A4','No. Inven y Descripcion'],
                   ['B4','PICTOGRAMAS'],
                   ['C4','INDICACION'],
                   ['D4','ANTES DE USAR...']
                   ]

        # COMBINAR CELDAS PARTE SUPERIOR
        self.ws.merge_cells("B1:C1")
        self.ws.merge_cells("C2:D2")
        self.ws.merge_cells("A3:B3")
        self.ws.merge_cells("C3:D3")
        # PICTOGRAMAS
        letra_picto = ['a','c','i','m','t']
        images = {}
        index = 0
        for filename in natsorted(glob.glob('C:/Users/Fernando.Lopez/Downloads/los_codigos/los_codigos/openpyxl/pruebas/*.png')):    
            images[letra_picto[index]] = filename
            index += 1
        for i,j in enumerate(parte_superior): # CICLO FOR: AQUI SE CARGA A "PARTE SUPERIOR" AL ARCHIVO
            self.ws[j[0]].value = j[1] # SE AÑADE EL TEXTO 
            self.ws[j[0]].alignment = Alignment(wrap_text= True)# SE ENCUADRA EL TEXTO
        # DESEA CARGAR LA INFORMACION?
        print('DESEA CARGAR LA INFORMACION?')
        op1 = input('>>> ')
        if op1 == 'si':# SI DESEO CARGAR LA INFORMACION
            print('Se cargaran los paquetes de datos')
            print('CUAL DESEA AÑADIR?')
            for i,producto in enumerate(self.lista_productos):# SE ITERA A TRAVES DE DICCIONARIO 'PRODUCTOS Y REACTIVOS' DE LA BASE DE DATOS  
                print(f'{i}: {producto}')
            print('elija segun su NOMBRE:')
            op2 = input('>>> ')
            # SE SELECCIONA EL PAQUETE DE  INFORMACIÓN
            if op2 in self.lista_productos:# SE COMPRUEBA SI EL STRING INGRESADO ESTA EN EL DICCIONARIO
                self.reactivos_hoja = []
                fila = 5# FILA DONDE SE EMPIEZA A AGREGAR LA INFORMACION
                print('Si esta')
                unidades = len(self.lista_productos[str(op2)].keys())-2# ES EL NUMERO DE REACTIVOS DEL PRODUCTO ELEGIDO
                print(f'reactivos: {unidades}')
                for i,reactivo in enumerate(self.lista_productos[str(op2)].keys()):# SE ACOMODAN LOS DATOS
                    print(reactivo)
                    
                    # AQUI ES CUANDO SE CARGA LA INFORMACION DE LOS REACTIVOS
                    
                    if i != len(self.lista_productos[str(op2)].keys())-1 and i != len(self.lista_productos[str(op2)].keys())-2:
                        # NO SE ELIGE EL ULTIMO NI EL PENULTIMO ELEMENTO DE CADA PRODUCTO
                        print(f'{i}# Iteracion ')
                        self.reactivos_hoja.append(['A'+str(fila),reactivo])# No inventario
                        descripcion = self.lista_productos[op2][reactivo]['descripcion']# DESCRIPCION
                        self.reactivos_hoja.append(['A'+str(fila+1),descripcion])
                        self.ws.merge_cells('A'+str(fila+1)+':A'+str(fila+2))# se combinan
                        peligro = self.lista_productos[op2][reactivo]['peligro']
                        if peligro != '0':
                            for k,pictograma in enumerate(peligro):
                                ruta = images[pictograma]
                                print(ruta)
                                self.ws.add_image(Image(ruta),anchor=chr(66+k)+str(fila))
                                
                                
                        fila += 3
                print(self.reactivos_hoja)
                for i,anadir in enumerate(self.reactivos_hoja):# SE AÑADEN AL EXCEL
                    self.ws[anadir[0]].value = anadir[1]
                    self.ws[anadir[0]].alignment = Alignment(wrap_text= True)# SE ENCUADRA EL TEXTO
                self.reactivos_hoja =  []
            else:# EL USUARIO SE EQUIVOCO Y NO ESTA EN EL DICCIONARIO
                print('no está')

            # AQUI SE CARGA LA ULTIMA PARTE DEL PROGRAMA
            
            
        else: # NO DESEO CARGAR LA INFORMACION
            pass
        

##      ULTIMA INSTRUCCION ANTES DE CERRAR EL ARCHIVO
        self.wb.save(f'{nombre}.xlsx')
        

### INICIALIZAMOS EL PROGRAMA ###
clase_save = Save_module()
clase_save.active_module()
